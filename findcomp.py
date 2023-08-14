# import pandas as pd
import asyncio, sys
import logging
import csv
import argparse
import openpyxl
from openpyxl.worksheet.cell_range import CellRange

from rep import Reputation, serialize
from config import CONFIG

#==============================================================================#

NL = '\n'
DESC = \
"""
Варианты параметров:
1) findcomp -i "inn1, inn2, inn3, ..." [-o "path/to/output.csv"]                                 - поиск компаний по списку ИНН, результаты в CSV файл
2) findcomp -x "source/excel.xlsx" -r "a1:a100" [--offset 1] [-o "path/to/output.csv"]           - поиск компаний по ИНН из Excel, результаты опционально в тот же файл (+ опционально в CSV файл)

Результат будет содержать следующие данные (строго в указанном порядке):
1) Краткое имя
2) Полное имя
3) Адрес
4) Руководитель
5) Сфера деятельности (основная по ОГРН)
6) Дата регистрации
7) Вебсайт (основной)
8) Тел. (список)
9) Email (список)

Если указан параметр --offset, то результаты будут записаны в исходный Excel-файл, начиная с указанного смещения по столбцам.

Если указан параметр --out, то результаты будут записаны в указанный файл в формате CSV (с заголовком).
"""

#==============================================================================#

async def main():
    parser = argparse.ArgumentParser(description=DESC, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('-i', '--inn', help='Перечень ИНН через запятую')
    parser.add_argument('-x', '--xl', nargs='?', help='Путь к исходному файлу Excel')
    parser.add_argument('-s', '--sheet', help='Лист Excel (название или индекс)')
    parser.add_argument('-r', '--rginn', help='Диапазон в Excel в котором расположены ИНН в столбец (напр. A1:A100)')
    parser.add_argument('--offset', type=int, help='Смещение по столбцам для результатов (по умолчанию не задано)')
    parser.add_argument('-m', '--max', type=int, help='Максимальное количество элементов для полей "Телефоны" и "Email"')
    parser.add_argument('-o', '--out', nargs='?', help='Путь к файлу для сохранения результатов в формате CSV')
    args = vars(parser.parse_args())

    # print(args)
    
    try:
        entities = []
        wb = None

        if args.get('inn', None):
            entities = [e.strip() for e in args['inn'].split(',')]
        elif args.get('xl', None):
            if not 'rginn' in args:
                raise argparse.ArgumentError(message='Для источника Excel надо указать исходный диапазон в параметре "--rginn"!')
            wb = openpyxl.load_workbook(filename=args['xl'])
            sheet = args.get('sheet', 0) or 0
            ws = wb.worksheets[sheet]
            wr = CellRange(f"{ws.title}!{args['rginn']}")
            entities = [str(c[0]) for c in ws.iter_rows(min_row=wr.min_row, max_row=wr.max_row, 
                                                        min_col=wr.max_col, max_col=wr.max_col, values_only=True)]                
        else:
            raise argparse.ArgumentError(message='В качестве источника укажите или ИНН через запятую в параметре')
        
        logging.info(f'Исходный список ИНН ({len(entities)}):{NL}{", ".join(entities)}')

        if not entities:
            raise Exception('Список ИНН организаций пуст!')

        async with Reputation(CONFIG.api_key.get_secret_value(), args.get('max', None)) as rep_:
            found_entities = await rep_.batch__search_entity_by_inn(entities)

        if not found_entities:
            raise Exception('Не найдена ни одна организация в списке!')
        
        offset = args.get('offset', None)
        if offset:
            logging.debug(f'Запись результатов в исходный Excel ("{args["xl"]}")...')
            for c in ws.iter_rows(min_row=wr.min_row, max_row=wr.max_row, min_col=wr.max_col, max_col=wr.max_col):
                cc = c[0]
                for e in found_entities:
                    if e.inn == str(cc.value):
                        ws.cell(row=cc.row, column=(cc.column + offset), value=e.name or '')
                        ws.cell(row=cc.row, column=(cc.column + offset + 1), value=e.full_name or '')
                        ws.cell(row=cc.row, column=(cc.column + offset + 2), value=e.address or '')
                        ws.cell(row=cc.row, column=(cc.column + offset + 3), value=e.manager or '')
                        ws.cell(row=cc.row, column=(cc.column + offset + 4), value=e.activity or '')
                        ws.cell(row=cc.row, column=(cc.column + offset + 5), value=e.date_registered or '')
                        ws.cell(row=cc.row, column=(cc.column + offset + 6), value=e.website or '')
                        ws.cell(row=cc.row, column=(cc.column + offset + 7), value='\n'.join(e.phones).strip())
                        ws.cell(row=cc.row, column=(cc.column + offset + 8), value='\n'.join(e.emails).strip())
                        break
            wb.save(args['xl'])
            logging.debug('Результаты сохранены в исходный Excel')

        outfile = args.get('out', None)
        if outfile:
            logging.debug(f'Запись результатов в CSV ("{outfile}")...')
            with open(outfile, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, dialect='excel')
                writer.writerow(['Краткое имя', 'Полное имя', 'Адрес', 'Руководитель',
                                 'Сфера деятельности', 'Дата регистрации', 'Вебсайт', 'Тел.', 'Email'])
                for e in found_entities:
                    writer.writerow([e.name or '', e.full_name or '', e.address or '', e.manager or '',
                                     e.activity or '', e.date_registered or '', e.website or '',
                                     '; '.join(e.phones).strip(), '; '.join(e.emails).strip()])
            logging.debug('Результаты сохранены в CSV')
        
        logging.info(f'Поиск завершен, найдены результаты по {len(found_entities)} организациям')
    
    except Exception as err:
        logging.exception(err)
        parser.print_help()
              

#==============================================================================#

if __name__ == '__main__':
    asyncio.run(main())